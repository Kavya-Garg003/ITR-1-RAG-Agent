"""
ITR-1 Excel Form Filler
========================
Takes parsed form data (from the agent pipeline), fills the official macro-enabled
ITR-1 government template using openpyxl, and combines all the key sheets
visually into a single consolidated worksheet.
"""

from __future__ import annotations
import uuid
import warnings
import shutil
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import range_boundaries

warnings.filterwarnings("ignore")

# Paths
PROJECT_ROOT  = Path(__file__).parent.parent
TEMPLATE_XLSM = PROJECT_ROOT / "knowledge-base" / "form_files" / "ITR1_AY_25-26_V1.7.xlsm"
OUTPUT_DIR    = PROJECT_ROOT / "agent-orchestrator" / "filled_forms"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CELL_MAP = [
    # ── Part A Personal Info ──────────────────────────────────────────────────
    ("Income Details", "C10",  "personal_info.first_name"),
    ("Income Details", "C11",  "personal_info.middle_name"),
    ("Income Details", "C12",  "personal_info.last_name"),
    ("Income Details", "C13",  "personal_info.pan"),
    ("Income Details", "C18",  "personal_info.address_flat"),
    ("Income Details", "C19",  "personal_info.address_premises"),
    ("Income Details", "C20",  "personal_info.address_street"),
    ("Income Details", "C21",  "personal_info.address_locality"),
    ("Income Details", "C22",  "personal_info.address_city"),
    ("Income Details", "C23",  "personal_info.address_state"),
    ("Income Details", "C24",  "personal_info.address_pin"),
    ("Income Details", "C25",  "personal_info.aadhaar"),
    ("Income Details", "C27",  "personal_info.mobile"),
    ("Income Details", "C28",  "personal_info.email"),
    ("Income Details", "AO25", "personal_info.aadhaar"),

    # ── Schedule S (Salary) ────────────────────────────────────────────────────
    ("Income Details", "AO39", "salary_income.salary_as_per_17_1"),
    ("Income Details", "AO40", "salary_income.perquisites_17_2"),
    ("Income Details", "AO41", "salary_income.profits_17_3"),
    ("Income Details", "AO42", "salary_income.gross_salary"),
    ("Income Details", "AO48", "salary_income.total_exempt_allowances"),
    ("Income Details", "AO49", "salary_income.net_salary"),
    ("Income Details", "AO50", "salary_income.standard_deduction_16ia"),
    ("Income Details", "AO51", "salary_income.entertainment_allowance_16ii"),
    ("Income Details", "AO52", "salary_income.professional_tax_16iii"),
    ("Income Details", "AO53", "salary_income.total_sec16_deductions"),
    ("Income Details", "AO54", "salary_income.taxable_salary"),

    # ── Schedule HP (House Property) ───────────────────────────────────────────
    ("Income Details", "AO59", "house_property.total_income_hp"),

    # ── Schedule OS (Other Sources) ────────────────────────────────────────────
    ("Income Details", "AO76", "other_sources.total_other_sources"),

    # ── Part B ATI (Tax Computation) ──────────────────────────────────────────
    ("Part B ATI", "AO5",  "tax_computation.gross_total_income"),
    ("Part B ATI", "AO6",  "tax_computation.taxable_income"),
    ("Part B ATI", "AO9",  "tax_computation.tax_before_rebate"),
    ("Part B ATI", "AO11", "tax_computation.rebate_87a"),
    ("Part B ATI", "AO12", "tax_computation.tax_after_rebate"),
    ("Part B ATI", "AO14", "tax_computation.surcharge"),
    ("Part B ATI", "AO15", "tax_computation.health_education_cess"),
    ("Part B ATI", "AO16", "tax_computation.total_tax_liability"),

    # ── TDS Schedule TDS1 (employer TDS) ──────────────────────────────────────
    ("TDS", "E10",  "tds_details.0.employer_tan"),
    ("TDS", "N10",  "tds_details.0.employer_name"),
    ("TDS", "V10",  "tds_details.0.income_chargeable"),
    ("TDS", "Z10",  "tds_details.0.tds_deducted"),
    ("TDS", "AD10", "tds_details.0.tds_claimed"),

    # ── Taxes Paid and Verification ────────────────────────────────────────────
    ("Taxes Paid and Verification", "AO5",  "tax_computation.tds_deducted"),
    ("Taxes Paid and Verification", "AO7",  "tax_computation.total_taxes_paid"),
    ("Taxes Paid and Verification", "AO8",  "tax_computation.tax_payable"),
    ("Taxes Paid and Verification", "AO9",  "tax_computation.refund"),
    ("Taxes Paid and Verification", "AO12", "personal_info.bank_account_number"),
    ("Taxes Paid and Verification", "AO13", "personal_info.bank_ifsc"),
]

def _get_nested(data: dict, dot_path: str) -> Any:
    parts = dot_path.split(".")
    cur = data
    for p in parts:
        if isinstance(cur, dict):
            cur = cur.get(p)
        elif isinstance(cur, list):
            try:
                cur = cur[int(p)]
            except (ValueError, IndexError):
                return None
        else:
            return None
        if cur is None:
            return None
    return cur

def _fmt(val: Any, is_numeric: bool = False) -> Any:
    if val is None:
        return 0 if is_numeric else ""
    if is_numeric:
        try:
            return round(float(val))
        except (ValueError, TypeError):
            return 0
    return str(val)

def write_cell_openpyxl(ws, cell_addr, value):
    try:
        cell = ws[cell_addr]
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell_addr in merged_range:
                    top_left_coord = merged_range.start_cell.coordinate
                    ws[top_left_coord] = value
                    return True
        else:
            ws[cell_addr] = value
            return True
    except Exception as e:
        print(f"Error writing cell {cell_addr} in sheet {ws.title}: {e}")
        return False

def copy_sheet_formatting_and_values(src_ws, dest_ws, start_row):
    # Copy column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width:
            dest_ws.column_dimensions[col_letter].width = max(
                dest_ws.column_dimensions[col_letter].width or 0,
                col_dim.width
            )
            
    # Copy rows and styles
    max_r = src_ws.max_row
    max_c = src_ws.max_column
    
    for row_idx in range(1, max_r + 1):
        dest_row_idx = row_idx + start_row - 1
        if row_idx in src_ws.row_dimensions:
            if src_ws.row_dimensions[row_idx].height:
                dest_ws.row_dimensions[dest_row_idx].height = src_ws.row_dimensions[row_idx].height
                
        for col_idx in range(1, max_c + 1):
            cell = src_ws.cell(row=row_idx, column=col_idx)
            dest_cell = dest_ws.cell(row=dest_row_idx, column=col_idx, value=cell.value)
            
            if cell.has_style:
                if cell.font:
                    dest_cell.font = openpyxl.styles.Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                        underline=cell.font.underline,
                        strikethrough=cell.font.strikethrough
                    )
                if cell.fill:
                    dest_cell.fill = openpyxl.styles.PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
                if cell.border:
                    dest_cell.border = openpyxl.styles.Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                if cell.alignment:
                    dest_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent
                    )
                dest_cell.number_format = cell.number_format
                
    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        dest_ws.merge_cells(
            start_row=min_row + start_row - 1,
            start_column=min_col,
            end_row=max_row + start_row - 1,
            end_column=max_col
        )
        
    return start_row + max_r + 2

def fill_itr1_excel(itr_data: dict, session_id: str) -> Path:
    unique_id = uuid.uuid4().hex[:8]
    out_path = OUTPUT_DIR / f"ITR1_filled_{session_id}_{unique_id}.xlsm"
    
    wb = openpyxl.load_workbook(str(TEMPLATE_XLSM), data_only=False, keep_vba=True)
    form = itr_data.get("itr1_form", itr_data)
    
    # 1. Fill values in the original sheets
    filled_count = 0
    for sheet_name, cell_addr, dot_path in CELL_MAP:
        val = _get_nested(form, dot_path)
        numeric_keys = {
            "salary", "income", "tax", "deduction", "tds", "rebate",
            "cess", "surcharge", "refund", "total", "gross", "net",
            "allowance", "exempt", "interest", "pension", "dividend",
        }
        is_num = any(k in dot_path for k in numeric_keys)
        formatted = _fmt(val, is_numeric=is_num)
        
        if formatted == 0 or formatted == "":
            continue
            
        ws = wb[sheet_name]
        if write_cell_openpyxl(ws, cell_addr, formatted):
            filled_count += 1
            
    # Extra fields
    first = _get_nested(form, "personal_info.first_name") or ""
    last  = _get_nested(form, "personal_info.last_name")  or ""
    full_name = f"{first} {last}".strip()
    if full_name:
        write_cell_openpyxl(wb["Taxes Paid and Verification"], "AO2", full_name)
        
    pan = _get_nested(form, "personal_info.pan") or ""
    if pan:
        write_cell_openpyxl(wb["Taxes Paid and Verification"], "AO3", pan)
        
    sec_80c = _get_nested(form, "deductions.sec_80c") or 0
    if sec_80c and float(sec_80c) > 0:
        write_cell_openpyxl(wb["80C"], "D9", "Various (from Form 16)")
        write_cell_openpyxl(wb["80C"], "H9", round(float(sec_80c)))
        
    sec_80d = _get_nested(form, "deductions.sec_80d") or 0
    if sec_80d and float(sec_80d) > 0:
        write_cell_openpyxl(wb["80D"], "H5", round(float(sec_80d)))
        
    sbi  = _get_nested(form, "other_sources.savings_bank_interest") or 0
    fd   = _get_nested(form, "other_sources.fd_interest")           or 0
    div  = _get_nested(form, "other_sources.dividends")             or 0
    
    os_items = []
    if sbi and float(sbi) > 0: os_items.append(("Interest from Savings Bank", round(float(sbi))))
    if fd and float(fd) > 0:   os_items.append(("Interest from Deposits", round(float(fd))))
    if div and float(div) > 0: os_items.append(("Dividend Income", round(float(div))))
    
    rows = [68, 69, 70, 71]
    for i, (nature, amt) in enumerate(os_items[:4]):
        write_cell_openpyxl(wb["Income Details"], f"J{rows[i]}", nature)
        write_cell_openpyxl(wb["Income Details"], f"AO{rows[i]}", amt)
        filled_count += 2
        
    # 2. Combine all sheets visually into a single worksheet
    combined_ws = wb.create_sheet(title="Combined ITR-1")
    
    core_sheets = ["Income Details", "80C", "80D", "TDS", "Taxes Paid and Verification", "Part B ATI"]
    start_row = 1
    
    for sh_name in core_sheets:
        if sh_name in wb.sheetnames:
            start_row = copy_sheet_formatting_and_values(wb[sh_name], combined_ws, start_row)
            
    # Remove all other sheets except the combined one
    for sh_name in list(wb.sheetnames):
        if sh_name != "Combined ITR-1":
            wb.remove(wb[sh_name])
            
    wb.save(str(out_path))
    print(f"[ITR1Filler] Saved combined official format Excel to {out_path}")
    return out_path

def get_filled_form_path(session_id: str) -> Path | None:
    p = OUTPUT_DIR / f"ITR1_filled_{session_id}.xlsx"
    return p if p.exists() else None

def export_to_pdf_win32(excel_path: Path) -> Path:
    import win32com.client
    import pythoncom
    
    pdf_path = excel_path.with_suffix(".pdf")
    if pdf_path.exists():
        pdf_path.unlink()
        
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AutomationSecurity = 3
        
        wb = excel.Workbooks.Open(str(excel_path.resolve()))
        wb.ActiveSheet.ExportAsFixedFormat(0, str(pdf_path.resolve()))
        wb.Close(SaveChanges=False)
    except Exception as e:
        print(f"Failed to export PDF via win32com: {e}")
        raise
    finally:
        if excel:
            excel.Quit()
        pythoncom.CoUninitialize()
        
    return pdf_path
